"""
Generador de Excel con diseño mejorado usando openpyxl
"""

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# Estilos
BLUE_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
GRAY_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

WHITE_FONT = Font(color="FFFFFF", bold=True, size=14)
BOLD_FONT = Font(bold=True)
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=12)
TOTAL_FONT = Font(bold=True, size=12)

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
WRAP_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
WRAP_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_RIGHT_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True)

# Anchos mínimos y máximos de columna
MIN_COL_WIDTH = 8
MAX_COL_WIDTH = 50


def generate_excel(data: dict[str, Any], invoice_index: int = 0, lang: str = "es") -> bytes:
    """
    Genera un archivo Excel con diseño profesional para una factura.

    Args:
        data: Datos parseados de la factura (formato del parser frontend)
        invoice_index: Índice de la factura (para lotes)
        lang: Idioma ('es' o 'en')

    Returns:
        Contenido del archivo Excel como bytes
    """
    # Traducciones
    t = TRANSLATIONS.get(lang, TRANSLATIONS["es"])

    invoice = data["invoices"][invoice_index]
    seller = data.get("seller", {})
    buyer = data.get("buyer", {})
    currency = data.get("fileHeader", {}).get("currencyCode", "EUR")

    wb = Workbook()
    ws = wb.active
    ws.title = t["sheet_name"]

    row = 1

    # === TÍTULO ===
    invoice_number = f"{invoice.get('series', '')}{invoice.get('series') and '/' or ''}{invoice.get('number', '')}"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    title_cell = ws.cell(row=row, column=1, value=f"{t['invoice']} {invoice_number}")
    title_cell.font = WHITE_FONT
    title_cell.fill = BLUE_FILL
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[row].height = 30
    row += 1

    # === DATOS BÁSICOS ===
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    info_text = f"{t['date']}: {invoice.get('issueDate', '')}    |    {t['currency']}: {currency}    |    Facturae: {data.get('version', '')}"
    info_cell = ws.cell(row=row, column=1, value=info_text)
    info_cell.fill = LIGHT_GRAY_FILL
    info_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[row].height = 22
    row += 2

    # === EMISOR Y RECEPTOR ===
    # Headers
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    seller_header = ws.cell(row=row, column=1, value=t["seller"])
    buyer_header = ws.cell(row=row, column=4, value=t["buyer"])
    for cell in [seller_header, buyer_header]:
        cell.font = TITLE_FONT
        cell.fill = GRAY_FILL
        cell.alignment = CENTER_ALIGN
    row += 1

    # Datos emisor/receptor
    party_data = [
        (t["name"], seller.get("name", ""), buyer.get("name", "")),
        (t["tax_id"], seller.get("taxId", ""), buyer.get("taxId", "")),
        (t["address"], _format_address(seller.get("address")), _format_address(buyer.get("address"))),
    ]

    for label, seller_val, buyer_val in party_data:
        label_cell1 = ws.cell(row=row, column=1, value=label)
        label_cell1.font = BOLD_FONT
        label_cell1.alignment = WRAP_ALIGN
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        value_cell1 = ws.cell(row=row, column=2, value=seller_val)
        value_cell1.alignment = WRAP_ALIGN
        label_cell2 = ws.cell(row=row, column=4, value=label)
        label_cell2.font = BOLD_FONT
        label_cell2.alignment = WRAP_ALIGN
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        value_cell2 = ws.cell(row=row, column=5, value=buyer_val)
        value_cell2.alignment = WRAP_ALIGN
        row += 1

    row += 1

    # === LÍNEAS DE DETALLE ===
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    lines_header = ws.cell(row=row, column=1, value=t["lines"])
    lines_header.font = TITLE_FONT
    lines_header.fill = GRAY_FILL
    lines_header.alignment = CENTER_ALIGN
    row += 1

    # Cabecera tabla
    line_headers = [t["line_num"], t["description"], t["quantity"], t["unit_price"], t["vat"], t["amount"]]

    for col, header in enumerate(line_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = LIGHT_GRAY_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
    row += 1

    # Filas de líneas
    lines = invoice.get("lines", [])
    for idx, line in enumerate(lines, start=1):
        ws.cell(row=row, column=1, value=idx).alignment = CENTER_ALIGN
        desc_cell = ws.cell(row=row, column=2, value=line.get("description", ""))
        desc_cell.alignment = WRAP_ALIGN
        ws.cell(row=row, column=3, value=line.get("quantity", 0)).alignment = RIGHT_ALIGN
        ws.cell(row=row, column=4, value=line.get("unitPrice", 0)).alignment = RIGHT_ALIGN
        ws.cell(row=row, column=5, value=f"{line.get('taxRate', 0)}%").alignment = CENTER_ALIGN
        ws.cell(row=row, column=6, value=line.get("grossAmount") or line.get("totalAmount", 0)).alignment = RIGHT_ALIGN

        # Aplicar bordes
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER

        # Formato número para columnas numéricas
        ws.cell(row=row, column=3).number_format = "#,##0.00"
        ws.cell(row=row, column=4).number_format = "#,##0.00"
        ws.cell(row=row, column=6).number_format = "#,##0.00"
        row += 1

    row += 1

    # === IMPUESTOS Y TOTALES ===
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    totals_header = ws.cell(row=row, column=1, value=t["totals"])
    totals_header.font = TITLE_FONT
    totals_header.fill = GRAY_FILL
    totals_header.alignment = CENTER_ALIGN
    row += 1

    totals = invoice.get("totals", {})
    taxes = invoice.get("taxes", [])

    # Desglose de impuestos
    for tax in taxes:
        tax_type = _get_tax_type_label(tax.get("type", ""), lang)
        tax_label = f"{tax_type} {tax.get('rate', 0)}%"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value=f"  {t['tax_base']}: {_format_currency(tax.get('base', 0), currency)}")
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row=row, column=5, value=f"{tax_label}: {_format_currency(tax.get('amount', 0), currency)}").alignment = RIGHT_ALIGN
        row += 1

    # Retenciones
    if totals.get("taxesWithheld", 0) != 0:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value=f"  {t['withholdings']}")
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row=row, column=5, value=_format_currency(-abs(totals.get("taxesWithheld", 0)), currency)).alignment = RIGHT_ALIGN
        row += 1

    row += 1

    # Total a pagar
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    total_label = ws.cell(row=row, column=1, value=t["total_to_pay"])
    total_label.font = TOTAL_FONT
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    total_value = ws.cell(row=row, column=5, value=_format_currency(totals.get("totalToPay", 0), currency))
    total_value.font = TOTAL_FONT
    total_value.alignment = RIGHT_ALIGN
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = LIGHT_GRAY_FILL
    row += 2

    # === INFORMACIÓN DE PAGO ===
    payment = invoice.get("payment", {})
    if payment:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        payment_header = ws.cell(row=row, column=1, value=t["payment"])
        payment_header.font = TITLE_FONT
        payment_header.fill = GRAY_FILL
        payment_header.alignment = CENTER_ALIGN
        row += 1

        if payment.get("paymentMeans"):
            label_cell = ws.cell(row=row, column=1, value=t["payment_method"])
            label_cell.font = BOLD_FONT
            label_cell.alignment = WRAP_ALIGN
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws.cell(row=row, column=2, value=_get_payment_means_label(payment.get("paymentMeans"), lang))
            row += 1

        if payment.get("dueDate"):
            label_cell = ws.cell(row=row, column=1, value=t["due_date"])
            label_cell.font = BOLD_FONT
            label_cell.alignment = WRAP_ALIGN
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws.cell(row=row, column=2, value=payment.get("dueDate"))
            row += 1

        if payment.get("iban"):
            label_cell = ws.cell(row=row, column=1, value="IBAN")
            label_cell.font = BOLD_FONT
            label_cell.alignment = WRAP_ALIGN
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws.cell(row=row, column=2, value=payment.get("iban"))
            row += 1

        if payment.get("bic"):
            label_cell = ws.cell(row=row, column=1, value="BIC")
            label_cell.font = BOLD_FONT
            label_cell.alignment = WRAP_ALIGN
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws.cell(row=row, column=2, value=payment.get("bic"))

    # Auto-ajustar anchos de columna
    _auto_fit_columns(ws)

    # Guardar a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _format_address(address: dict | None) -> str:
    """Formatea una dirección como string"""
    if not address:
        return ""
    parts = [
        address.get("street", ""),
        address.get("postCode", ""),
        address.get("town", ""),
        address.get("province", ""),
    ]
    return ", ".join(p for p in parts if p)


def _format_currency(amount: float, currency: str) -> str:
    """Formatea un importe con moneda"""
    return f"{amount:,.2f} {currency}"


def _get_tax_type_label(code: str, lang: str) -> str:
    """Obtiene la etiqueta del tipo de impuesto"""
    labels = {
        "es": {"01": "IVA", "04": "IRPF"},
        "en": {"01": "VAT", "04": "Income Tax"},
    }
    return labels.get(lang, labels["es"]).get(code, code)


def _get_payment_means_label(code: str, lang: str) -> str:
    """Obtiene la etiqueta del método de pago"""
    labels = {
        "es": {
            "01": "Al contado",
            "02": "Recibo domiciliado",
            "03": "Recibo",
            "04": "Transferencia",
            "05": "Letra aceptada",
            "06": "Crédito documentario",
            "07": "Contrato adjudicación",
            "08": "Letra de cambio",
            "09": "Pagaré a la orden",
            "10": "Pagaré no a la orden",
            "11": "Cheque",
            "12": "Reposición",
            "13": "Especiales",
            "14": "Compensación",
            "15": "Giro postal",
            "16": "Cheque conformado",
            "17": "Cheque bancario",
            "18": "Pago contra reembolso",
            "19": "Pago mediante tarjeta",
        },
        "en": {
            "01": "Cash",
            "02": "Direct debit",
            "03": "Receipt",
            "04": "Bank transfer",
            "05": "Accepted bill",
            "06": "Documentary credit",
            "07": "Award contract",
            "08": "Bill of exchange",
            "09": "Promissory note to order",
            "10": "Promissory note not to order",
            "11": "Check",
            "12": "Replacement",
            "13": "Special",
            "14": "Compensation",
            "15": "Postal order",
            "16": "Certified check",
            "17": "Bank check",
            "18": "Cash on delivery",
            "19": "Card payment",
        },
    }
    return labels.get(lang, labels["es"]).get(code, code)


def _auto_fit_columns(ws: Worksheet) -> None:
    """
    Auto-ajusta el ancho de las columnas basándose en el contenido.
    Tiene en cuenta celdas combinadas y aplica límites min/max.
    """
    column_widths: dict[int, float] = {}

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            col_idx = cell.column

            # Calcular ancho necesario para este valor
            value_str = str(cell.value)
            # Considerar saltos de línea (tomar la línea más larga)
            lines = value_str.split("\n")
            max_line_len = max(len(line) for line in lines)

            # Añadir padding y ajustar por tipo de fuente
            width = max_line_len + 2
            if cell.font and cell.font.bold:
                width *= 1.1  # Las negritas ocupan más espacio

            # Guardar el máximo para esta columna
            if col_idx not in column_widths:
                column_widths[col_idx] = MIN_COL_WIDTH
            column_widths[col_idx] = max(column_widths[col_idx], width)

    # Anchos mínimos específicos por columna (para labels y descripciones)
    min_widths_by_col = {
        1: 16,  # Columna de labels (Nombre, Dirección, etc.)
        2: 35,  # Columna de valores/descripciones
        4: 14,  # Columna de labels receptor
        5: 30,  # Columna de valores receptor
    }

    # Aplicar anchos con límites
    for col_idx, width in column_widths.items():
        # Usar mínimo específico si existe, sino el global
        col_min = min_widths_by_col.get(col_idx, MIN_COL_WIDTH)
        final_width = max(col_min, min(MAX_COL_WIDTH, width))
        ws.column_dimensions[get_column_letter(col_idx)].width = final_width


TRANSLATIONS = {
    "es": {
        "sheet_name": "Factura",
        "invoice": "FACTURA",
        "date": "Fecha",
        "currency": "Moneda",
        "seller": "EMISOR",
        "buyer": "RECEPTOR",
        "name": "Nombre",
        "tax_id": "NIF/CIF",
        "address": "Dirección",
        "lines": "DETALLE",
        "line_num": "Nº",
        "description": "Descripción",
        "quantity": "Cantidad",
        "unit_price": "Precio unit.",
        "vat": "IVA",
        "amount": "Importe",
        "totals": "TOTALES",
        "tax_base": "Base imponible",
        "withholdings": "Retenciones",
        "total_to_pay": "TOTAL A PAGAR",
        "payment": "INFORMACIÓN DE PAGO",
        "payment_method": "Forma de pago",
        "due_date": "Vencimiento",
    },
    "en": {
        "sheet_name": "Invoice",
        "invoice": "INVOICE",
        "date": "Date",
        "currency": "Currency",
        "seller": "SELLER",
        "buyer": "BUYER",
        "name": "Name",
        "tax_id": "Tax ID",
        "address": "Address",
        "lines": "LINE ITEMS",
        "line_num": "#",
        "description": "Description",
        "quantity": "Quantity",
        "unit_price": "Unit price",
        "vat": "VAT",
        "amount": "Amount",
        "totals": "TOTALS",
        "tax_base": "Tax base",
        "withholdings": "Withholdings",
        "total_to_pay": "TOTAL TO PAY",
        "payment": "PAYMENT INFORMATION",
        "payment_method": "Payment method",
        "due_date": "Due date",
    },
}
